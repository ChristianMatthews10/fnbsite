


import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image

# Create a new workbook and add the input sheet
workbook = openpyxl.Workbook()
input_sheet = workbook.active
input_sheet.title = "Inputs"

# Add headings to the input sheet
input_sheet.cell(row=1, column=1, value="Claim reference")
input_sheet.cell(row=2, column=1, value="Commercial invoice amount")
input_sheet.cell(row=3, column=1, value="Repair cost")
input_sheet.cell(row=4, column=1, value="Shipping costs")
input_sheet.cell(row=5, column=1, value="Packaging costs")
input_sheet.cell(row=6, column=1, value="Currency")
input_sheet.cell(row=7, column=1, value="Exchange rate")
input_sheet.cell(row=8, column=1, value="VAT rate")
input_sheet.cell(row=9, column=1, value="Image attachment")

# Add cells for user input
for row in range(1, 10):
    input_sheet.cell(row=row, column=2).alignment = Alignment(horizontal='center')
    input_sheet.cell(row=row, column=2).font = Font(bold=True)

# Add a cell for the image attachment
input_sheet.cell(row=9, column=2).value = "Insert image here"

# Add the output sheet
output_sheet = workbook.create_sheet(title="Outputs")

# Add headings to the output sheet
output_sheet.cell(row=1, column=1, value="Claim reference")
output_sheet.cell(row=1, column=2, value="Commercial invoice amount")
output_sheet.cell(row=1, column=3, value="Repair cost")
output_sheet.cell(row=1, column=4, value="Shipping costs")
output_sheet.cell(row=1, column=5, value="Packaging costs")
output_sheet.cell(row=1, column=6, value="Currency")
output_sheet.cell(row=1, column=7, value="Exchange rate")
output_sheet.cell(row=1, column=8, value="VAT rate")
output_sheet.cell(row=1, column=9, value="Image attachment")
output_sheet.cell(row=1, column=10, value="Excess")
output_sheet.cell(row=1, column=11, value="Total")
output_sheet.cell(row=1, column=12, value="VAT")
output_sheet.cell(row=1, column=13, value="Claim amount")

# Set the width of the columns on the output sheet
for col in range(1, 14):
    output_sheet.column_dimensions[get_column_letter(col)].width = 15

# Add the user input from the input sheet to the output sheet
for row in range(1, 10):
    output_sheet.cell(row=2, column=row, value=input_sheet.cell(row=row, column=2).value)

# Add the image attachment to the output sheet


# Calculate the values and fill them in the output sheet
commercial_invoice = float(input_sheet.cell(row=2, column=2).value)
repair_cost = float(input_sheet.cell(row=3, column=2).value)
shipping_cost = float(input_sheet.cell(row=4, column=2).value)
packaging_cost = float(input_sheet.cell(row=5, column=2).value)
exchange_rate = float(input_sheet.cell(row=7, column=2).value)
